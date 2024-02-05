import tkinter as tk
from tkinter import ttk, simpledialog, filedialog
import pandas as pd
from datetime import datetime, timedelta
import os
import json
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from tkinter import messagebox
import math
from openpyxl.styles import Font
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker

# Global Variables
running = False
cycle_start_time = None
current_stage_start_time = None
stage_times = []
user_data = {}
current_stage = 0
cycle_number = 1
delay_start_time = None
is_delay = False
delay_durations = []
delay_reasons = []
delay_timer_start_time = None
delay_elapsed_time = timedelta(0)
updating_delay_timer = False
cycles_data = []
cycle_stages = {
    "No Corners": ["FTram1", "Loading", "RTram1", "Dumping"],
    "1 Corner": ["FTram1", "Corner 1", "FTram2", "Loading", "RTram1", "Corner Rev 1", "RTram2", "Dumping"],
    "2 Corners": ["FTram1", "Corner 1", "FTram2", "Corner 2", "FTram3", "Loading", "RTram1", "Corner Rev 1", "RTram2", "Corner Rev 2", "RTram3", "Dumping"],
    "3 Corners": ["FTram1", "Corner 1", "FTram2", "Corner 2", "FTram3", "Corner 3", "FTram4", "Loading", "RTram1", "Corner Rev 1", "RTram2", "Corner Rev 2", "RTram3", "Corner Rev 3", "RTram4", "Dumping"]
}


# Stage names
stages = ["Tramming", "Loading", "Tramming Back", "Dumping", 
          "Corner 1", "Tramming", "Corner 2", "Tramming", "Corner 3"]

def mm_ss_formatter(x, pos):
    """Convert seconds to 'MM:SS' format."""
    minutes = int(x // 60)
    seconds = int(x % 60)
    return f"{minutes:02d}:{seconds:02d}"

# Function to update the stage label
def update_stage_label():
    cycle_type = cycle_type_combobox.get()
    current_stages = cycle_stages.get(cycle_type, ["Unknown"])
    
    if current_stage < len(current_stages):
        current_stage_label.config(text=f"Current Stage: {current_stages[current_stage]}")
    else:
        current_stage_label.config(text="Cycle Completed")

def save_cycle_data_to_json():
    with open('temporary_cycle_data.json', 'w') as file:
        json.dump(cycles_data, file, indent=4, default=str)

# Timer Functionality
def update_timer():
    if running:
        now = datetime.now()
        cycle_elapsed = now - cycle_start_time
        stage_elapsed = now - current_stage_start_time

        # Format elapsed time as whole seconds
        cycle_elapsed_seconds = int(cycle_elapsed.total_seconds())
        stage_elapsed_seconds = int(stage_elapsed.total_seconds())

        timer_label.config(text=f"Total Time: {cycle_elapsed_seconds}s\nStage Time: {stage_elapsed_seconds}s")
        window.after(1000, update_timer)

def update_delay_timer():
    global delay_timer_start_time, delay_elapsed_time, updating_delay_timer
    if updating_delay_timer:
        current_time = datetime.now()
        total_delay_time = (current_time - delay_timer_start_time) + delay_elapsed_time
        # Format the delay time and update the delay timer label
        delay_timer_label.config(text=f"Delay Time: {int(total_delay_time.total_seconds())}s")
        window.after(1000, update_delay_timer)

# Cycle Management Functions
def start_cycle():
    global cycle_start_time, current_stage_start_time, running, stage_times, current_stage, cycle_number
    cycle_start_time = datetime.now()
    current_stage_start_time = cycle_start_time
    running = True
    stage_times = []
    current_stage = 0
    cycle_number = 1
    update_timer()
    next_stage_button.config(state="normal")
    delay_button.config(state="normal")
    stop_cycle_button.config(state="normal")
    update_stage_label()
    cycle_counter_label.config(text=f"Cycle Count: {cycle_number}")




def next_stage(event=None):
    global current_stage_start_time, stage_times, current_stage, cycle_start_time, delay_durations, delay_reasons, cycle_number
    now = datetime.now()
    stage_duration = now - current_stage_start_time
    rounded_seconds = math.ceil(stage_duration.total_seconds())
    rounded_duration = timedelta(seconds=rounded_seconds)
    stage_times.append(rounded_duration)
    current_stage_start_time = now
    current_stage += 1
    update_stage_label()

    cycle_type = cycle_type_combobox.get()
    current_stages = cycle_stages.get(cycle_type, [])
    
    if current_stage >= len(cycle_stages[cycle_type_combobox.get()]):
        save_cycle()  # Save the current cycle
        # Start new cycle logic
        current_stage = 0
        stage_times = []
        cycle_start_time = datetime.now()
        current_stage_start_time = cycle_start_time
        update_timer()
        update_stage_label()
        cycle_counter_label.config(text=f"Cycle Count: {cycle_number}")

def save_cycle():
    global delay_durations, delay_reasons, stage_times, cycle_start_time, cycle_number, cycles_data, remote_state_combobox
    cycle_end_time = datetime.now()
    
    # Calculate total cycle duration as timedelta
    total_cycle_duration = cycle_end_time - cycle_start_time
    if delay_durations:
        total_cycle_duration += sum(delay_durations, timedelta(0))
    
    # Round the total duration to the nearest second
    total_seconds = total_cycle_duration.total_seconds()
    rounded_seconds = math.ceil(total_seconds)
    rounded_duration = timedelta(seconds=rounded_seconds)
    
    # Format the duration into a string as HH:MM:SS
    formatted_duration = str(rounded_duration)
    
    # Prepare cycle data dictionary
    cycle_data = {
        'cycle_type': cycle_type_combobox.get(),
        'cycle_number': cycle_number,
        'remote_state': remote_state_combobox.get(),
        'start_time': str(cycle_start_time),
        'end_time': str(cycle_end_time),
        'total_duration': formatted_duration,  # Use the formatted string here
        'delay_durations': [str(d) for d in delay_durations],  # Convert to string
        'delay_reasons': list(delay_reasons),
        'stage_times': [str(t) for t in stage_times]
    }

    cycles_data.append(cycle_data)
    save_cycle_data_to_json()  # Make sure this call is here
    cycle_number += 1
    stage_times.clear()  # Clear the stage times for the next cycle
    delay_durations.clear()  # Clear the delay data for the next cycle
    delay_reasons.clear()  # Clear the delay reasons for the next cycle

    load_cycle_numbers()

def start_delay():
    global delay_start_time, is_delay, running, delay_timer_start_time, updating_delay_timer, delay_elapsed_time
    if not is_delay:
        delay_start_time = datetime.now()
        is_delay = True
        running = False  # Stop the main timer updates
        delay_timer_start_time = datetime.now()
        updating_delay_timer = True
        delay_elapsed_time = timedelta(0)
        red_light.grid_remove()  # Hide red light
        green_light.grid()  # Show green light
        delay_running.grid()
        delay_timer_label.grid()
        update_delay_timer()  # Start updating the delay timer
    else:
        end_delay()

def end_delay():
    global delay_start_time, is_delay, delay_durations, delay_reasons, running, updating_delay_timer, delay_elapsed_time, delay_timer_start_time
    now = datetime.now()
    delay_duration = now - delay_start_time
    delay_durations.append(delay_duration)
    is_delay = False
    running = True  # Resume the main timer
    update_timer()
    updating_delay_timer = False
    # Define the reasons and show the dialog
    delay_reasons_list = ["Machine Fault", "Chair Fault", "ACS Fault", "Mapping Fault", 
                          "Operator Break", "Obstacle Detected", "Initialisation", "Occupants in Level"]
    dialog = DelayReasonDialog(window, delay_reasons_list)
    selected_reasons = dialog.reasons
    if selected_reasons:
        delay_reasons.extend(selected_reasons)

    delay_elapsed_time += datetime.now() - delay_timer_start_time

    green_light.grid_remove()  # Hide green light
    delay_running.grid_remove()
    delay_timer_label.config(text="Delay Time: 0s")
    delay_timer_label.grid_remove()
    red_light.grid()  # Show red light

class DelayReasonDialog(tk.Toplevel):
    def __init__(self, parent, delay_reasons_list):
        super().__init__(parent)
        self.delay_reasons_list = delay_reasons_list
        
        self.title("Delay Reason")
        self.geometry("300x300")  # Adjust size as needed
        self.iconbitmap(r'./icon/icon.ico')
        self.config(bg='darkgray')

        self.update_idletasks()  # Update the dialog to get its size
        x_center = int(parent.winfo_screenwidth() / 2 - self.winfo_width() / 2)
        y_center = int(parent.winfo_screenheight() / 2 - self.winfo_height() / 2)
        self.geometry(f"+{x_center}+{y_center}")  # Move the dialog to the center

        tk.Label(self, text="Select the reason(s) for the delay:").pack(pady=10)
        
        self.check_vars = []  # To store the variables associated with the checkboxes
        for reason in delay_reasons_list:
            var = tk.BooleanVar()
            chk = tk.Checkbutton(self, text=reason, variable=var)
            chk.pack(anchor='w')
            self.check_vars.append(var)

        submit_button = tk.Button(self, text="Submit", command=self.submit_reason)
        submit_button.pack(pady=10)

        self.reasons = []  # To store the selected reasons

        self.transient(parent)  # Set to be on top of the main window
        self.grab_set()  # Modal
        self.wait_window()

    def submit_reason(self):
        # Use the instance variable here
        self.reasons = [self.delay_reasons_list[i] for i, var in enumerate(self.check_vars) if var.get()]
        self.destroy()

def parse_timedelta(time_str):
    if not time_str:
        return timedelta(0)
    h, m, s = map(float, time_str.split(':'))
    return timedelta(hours=h, minutes=m, seconds=s)    

# Excel Export Function
def prepare_data_for_export(cycles_data):
    expanded_data = []
    for cycle in cycles_data:
        current_stages = cycle_stages.get(cycle['cycle_type'], ["Unknown"])
        delay_durations = [parse_timedelta(d) for d in cycle['delay_durations']]
        combined_delay_duration = str(sum(delay_durations, timedelta(0)))

        expanded_cycle = {
            'Date': date_entry.get(),
            'Site': site_entry.get(),
            'Production Area': production_area_entry.get(),
            'Loader': loader_entry.get(),
            'Remote State': cycle['remote_state'],
            'Operator': operator_entry.get(),
            'Cycle Type': cycle_type_combobox.get(),
            'Cycle': cycle['cycle_number'],
            'Start Time': cycle['start_time'],
            'End Time': cycle['end_time'],
            'Total Duration': cycle['total_duration'],
            'Combined Delay Duration': combined_delay_duration,
            'Combined Delay Reasons': '; '.join(cycle['delay_reasons']),
        }
        
        for i, stage_time in enumerate(cycle['stage_times']):
            if i < len(current_stages):
                stage_name = current_stages[i]
                expanded_cycle[stage_name + ' Time'] = stage_time
            else:
                # Handle unexpected extra stages
                expanded_cycle[f'Extra Stage {i+1} Time'] = stage_time

        expanded_data.append(expanded_cycle)

    return expanded_data


def save_to_excel(user_data, cycles_data, filename):
    global expanded_data
    # User data in a separate DataFrame
    user_df = pd.DataFrame([user_data])
    expanded_data = prepare_data_for_export(cycles_data)
    # Cycle data in another DataFrame
    cycle_df = pd.DataFrame(prepare_data_for_export(cycles_data))

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write user data
        user_df.to_excel(writer, index=False, sheet_name="Data", startrow=0, header=True)
        # Write cycle data
        cycle_df.to_excel(writer, index=False, sheet_name="Data", startrow=3)

        # Get the workbook and the sheet
        workbook = writer.book
        worksheet = writer.sheets['Data']

        # Autofit columns for the entire sheet
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                try: 
                    # Check the length of the cell content
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Adjust the width factor as needed
            worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width

        for idx, cycle in enumerate(expanded_data, start=1):  # start=1 because Excel rows start at 1
            if cycle['Cycle'] in flagged_cycles:
                for col in range(1, len(cycle) + 1):
                    cell = writer.sheets["Data"].cell(row=idx + 4, column=col)  # +3 if there are 3 rows before cycle data starts
                    cell.font = Font(color="FF0000")  # Apply red font color

    # Prompt user to save the file to a specific location
def save_as_prompt(original_filename):
    # Prompt the user to choose a location and filename to save the Excel file
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")], confirmoverwrite=False)
    if file_path:
        # Check if file exists and confirm overwrite if it does
        if os.path.exists(file_path):
            overwrite = tk.messagebox.askyesno("Overwrite File", "The file already exists. Do you want to overwrite it?")
            if not overwrite:
                return

        # Move the file to the new location, overwriting if necessary
        shutil.move(original_filename, file_path)
    else:
        # If no file path is provided, keep the original file
        print("File save cancelled; original file retained.")


def get_user_data():
    return {
        'Date': date_entry.get(),
        'Site': site_entry.get(),
        'Production Area': production_area_entry.get(),
        'Tramming Distance': tramming_distance_entry.get(),
        'Loader': loader_entry.get(),
        'Operator': operator_entry.get(),
        'Stope Size (Tonnes)': stope_size_entry.get(),
        'Stockpile Size': stockpile_size_entry.get(),
        'Mapping Time': mapping_time_entry.get(),
        'Number of Corners': number_of_corners_combobox.get(),
    }

def stop_cycle():
    global running, current_stage_start_time, stage_times, current_stage
    if running:
        now = datetime.now()
        stage_duration = now - current_stage_start_time

        # Ensure the list is long enough
        while len(stage_times) <= current_stage:
            stage_times.append(timedelta(0))

        # Update the stage time
        stage_times[current_stage] = stage_duration

        # Save the cycle if it's at the 'Dumping' stage
        if current_stage == stages.index("Dumping"):
            save_cycle()
            delay_durations.clear()
            delay_reasons.clear()
        
        running = False
        update_timer()  # Stop the timer

def load_cycle_numbers():
    try:
        with open('temporary_cycle_data.json', 'r') as file:
            cycles = json.load(file)
            cycle_listbox.delete(0, tk.END)  # Clear existing entries
            for cycle in cycles:
                # Format the string with cycle number and remote state
                listbox_entry = f"Cycle {cycle['cycle_number']} - {cycle['remote_state']}"
                cycle_listbox.insert(tk.END, listbox_entry)
    except FileNotFoundError:
        print("JSON file not found.")


def flag_cycle():
    selected_index = cycle_listbox.curselection()
    if selected_index:
        selected_entry = cycle_listbox.get(selected_index[0])
        # Extract the cycle number from the entry
        selected_cycle_number = int(selected_entry.split()[1])  # Assumes format "Cycle X - State"
        flagged_cycles.add(selected_cycle_number)


def save_and_clear():
    try:
        with open('temporary_cycle_data.json', 'r') as file:
            cycles_data = json.load(file)

        if cycles_data:
            user_data = get_user_data()
            filename = 'temporary_cycles_data.xlsx'
            save_to_excel(user_data, cycles_data, filename)
            save_as_prompt(filename)
            cycles_data.clear()  # Clear data for the next session
            os.remove('temporary_cycle_data.json')  # Clear the temporary file
        else:
            print("No cycle data to save.")
    except FileNotFoundError:
        print("No cycle data file found.")

def on_close():
    global user_data
    if messagebox.askyesno("Save Data", "Do you want to save the data before closing?"):
        # Assuming 'user_data' and 'cycles_data' are your data
        # You might need to adjust how you access these variables
        try:
            if cycles_data:  # Check if there is any cycle data to save
                filename = 'temporary_cycles_data.xlsx'  # Filename for the Excel file
                save_to_excel(user_data, cycles_data, filename)  # Corrected function call
                save_as_prompt(filename)  # Function to prompt user to save the file
        except NameError as e:
            print("Error:", e)
            print("No data available to save.")
    else:
        if os.path.exists('temporary_cycle_data.json'):
            os.remove('temporary_cycle_data.json')

    window.destroy()  # Close the application

def process_cycle_data():
    try:
        with open('temporary_cycle_data.json', 'r') as file:
            cycles = json.load(file)

        processed_data = {'Auto': [], 'Remote': [], 'Ass. Steer': []}
        min_cycle_numbers = {'Auto': float('inf'), 'Remote': float('inf'), 'Ass. Steer': float('inf')}

        for cycle in cycles:
            state = cycle['remote_state']
            # Splitting total_duration into main part and milliseconds
            total_time_str, _, ms_str = cycle['total_duration'].partition('.')
            total_time = datetime.strptime(total_time_str, '%H:%M:%S')
            ms = int(ms_str) if ms_str else 0
            total_seconds = total_time.hour * 3600 + total_time.minute * 60 + total_time.second + ms / 1000.0

            processed_data[state].append((cycle['cycle_number'], total_seconds))
            min_cycle_numbers[state] = min(min_cycle_numbers[state], cycle['cycle_number'])

        # Adjust cycle numbers
        for state in processed_data:
            processed_data[state] = [(cycle_num - min_cycle_numbers[state] + 1, time) for cycle_num, time in processed_data[state]]

        return processed_data
    except FileNotFoundError:
        print("JSON file not found.")
        return {}
    
def create_graph(processed_data):
    plt.figure(figsize=(10, 6))

    # Define colors for each state
    colors = {
        'Auto': 'blue',
        'Remote': 'gold',
        'Ass. Steer': 'white'
    }

    for state, data in processed_data.items():
        if data:  # Check if there is data for the state
            cycle_nums, times = zip(*data)
            plt.plot(cycle_nums, times, label=state, color=colors.get(state, 'black'))

    plt.xlabel('Cycle Number')
    plt.ylabel('Total Cycle Time (MM:SS)')
    plt.title('Total Cycle Time by Remote State')
    plt.legend()
    plt.grid(True)

    # Set Y-axis formatter
    plt.gca().yaxis.set_major_formatter(ticker.FuncFormatter(mm_ss_formatter))

    plt.show()

def show_graph():
    processed_data = process_cycle_data()
    create_graph(processed_data)


# GUI Application
window = tk.Tk()
window.protocol("WM_DELETE_WINDOW", on_close)
window.bind("<Right>", lambda event: next_stage() if running else None)


window.title("Autonomous Loader Cycle Recorder")
window.iconbitmap(r'./icon/icon.ico')

style = ttk.Style()

# Configure a new style for the Combobox
style.theme_use('clam')  # Use a theme that allows customizing
style.configure("BlackBorder.TCombobox", fieldbackground="white", background="white", bordercolor="black", borderwidth=1)

window.config(bg='darkgray')

tk.Label(text="Date:", bg='darkgrey').grid(row=0, column=0, sticky='e', padx=10, pady=(10, 5))
date_entry = tk.Entry(borderwidth=1, relief="solid")
date_entry.grid(row=0, column=1, sticky='w', pady=(10, 5))

tk.Label(text="Site:", bg='darkgrey').grid(row=1, column=0, sticky='e', padx=10, pady=5)
site_entry = tk.Entry(borderwidth=1, relief="solid")
site_entry.grid(row=1, column=1, sticky='w', pady=5)

tk.Label(text="Production Area:", bg='darkgrey').grid(row=0, column=2, sticky='w', pady=(10, 5))
production_area_entry = tk.Entry(borderwidth=1, relief="solid")
production_area_entry.grid(row=0, column=3, sticky='w', pady=(10, 5))

tk.Label(text="Tramming Distance:", bg='darkgrey').grid(row=1, column=2, sticky='w', pady=5)
tramming_distance_entry = tk.Entry(borderwidth=1, relief="solid")
tramming_distance_entry.grid(row=1, column=3, sticky='w', pady=5)

tk.Label(text="Loader:", bg='darkgrey').grid(row=0, column=4, sticky='w', pady=(10, 5))
loader_entry = tk.Entry(borderwidth=1, relief="solid")
loader_entry.grid(row=0, column=5, sticky='w', padx=10, pady=(10, 5))

tk.Label(text="Operator:", bg='darkgrey').grid(row=1, column=4, sticky='w', pady=5)
operator_entry = tk.Entry(borderwidth=1, relief="solid")
operator_entry.grid(row=1, column=5, sticky='w', padx=10, pady=5)

tk.Label(text="Stope Size (Tonnes):", bg='darkgrey').grid(row=2, column=0, sticky='e', padx=10, pady=5)
stope_size_entry = tk.Entry(borderwidth=1, relief="solid")
stope_size_entry.grid(row=2, column=1, sticky='w', pady=5)

tk.Label(text="Stockpile Size:", bg='darkgrey').grid(row=2, column=2, sticky='w', pady=5)
stockpile_size_entry = tk.Entry(borderwidth=1, relief="solid")
stockpile_size_entry.grid(row=2, column=3, sticky='w', pady=5)

tk.Label(text="Mapping Time:", bg='darkgrey').grid(row=2, column=4, sticky='w', pady=5)
mapping_time_entry = tk.Entry(borderwidth=1, relief="solid")
mapping_time_entry.grid(row=2, column=5, sticky='w', padx=10, pady=5)

# Dropdown for Number of Corners
tk.Label(text="Number of Corners:", bg='darkgrey').grid(row=3, column=0, sticky='e', padx=10, pady=5)
number_of_corners_combobox = ttk.Combobox(style="BlackBorder.TCombobox", values=["0", "1", "2", "3"])
number_of_corners_combobox.grid(row=3, column=1, sticky='e', pady=5)

tk.Label(text="Cycle Type:", bg='darkgrey').grid(row=3, column=2, sticky='w', pady=5)
cycle_type_combobox = ttk.Combobox(style="BlackBorder.TCombobox", values=["No Corners", "1 Corner", "2 Corners", "3 Corners"])
cycle_type_combobox.grid(row=3, column=3, sticky='e', pady=5)

tk.Label(text="Remote State:", bg='darkgrey').grid(row=3, column=4, sticky='w', pady=5)
remote_state_combobox = ttk.Combobox(style="BlackBorder.TCombobox", values=["Remote", "Ass. Steer", "Auto", "Mixed"])
remote_state_combobox.grid(row=3, column=5, sticky='e', padx=10, pady=5)

timer_label = tk.Label(text="Cycle Timer", bg='darkgrey', borderwidth=1, relief="solid")
timer_label.grid(row=7, column=0, sticky='nsew', padx=(10, 5), pady=10)

current_stage_label = tk.Label(width=20, text="Current Stage", bg='darkgrey', borderwidth=1, relief="solid")
current_stage_label.grid(row=9, column=0, sticky='nsew', padx=(10,5))

delay_timer_label = tk.Label(text="Delay Time: 0s", bg='darkgrey')
delay_timer_label.grid(row=9, column=3, sticky='nsew', padx=(0,10))
delay_timer_label.grid_remove()

start_button = tk.Button(text="Start Recording", bg='lightgray', command=start_cycle)
start_button.grid(row=7, column=1, sticky='nsew',padx=(0, 5), pady=10)

next_stage_button = tk.Button(text="Next Stage", bg='lightgray', command=next_stage, state="disabled")
next_stage_button.grid(row=9, column=1, sticky='nsew', padx=(0, 5))

delay_button = tk.Button(text="Start/Stop Delay", bg='lightgray', command=start_delay, state="disabled")
delay_button.grid(row=7, column=2, sticky='nsew', pady=10, padx=(0,10))

stop_cycle_button = tk.Button(text="Stop Recording", bg='lightgray', command=stop_cycle, state="disabled")
stop_cycle_button.grid(row=11, column=0, sticky='ew', padx=(10, 5), pady=10)

save_clear_button = tk.Button(text="Save & Clear", bg='lightgray', command=save_and_clear)
save_clear_button.grid(row=11, column=5, sticky='ew', padx=10, pady=10)

red_light = tk.Canvas(width=10, height=10, bg='darkgrey', highlightthickness=0)
red_light.create_oval(6, 6, 20, 20, fill='darkgrey')
red_light.grid(row=7, column=3, sticky='nesw', pady=10)

green_light = tk.Canvas(width=10, height=10, bg='darkgrey', highlightthickness=0)
green_light.create_oval(6, 6, 20, 20, fill='green')
green_light.grid(row=7, column=3, sticky='nesw', pady=10)
green_light.grid_remove()

delay_running = tk.Label(text="Delay Active!", bg='darkgrey', borderwidth=1, relief="solid")
delay_running.grid(row=7, column=3, sticky='w', padx=(25,0))
delay_running.grid_remove()

cycle_counter_label = tk.Label(width=10,text="Cycle Count: 0", bg='darkgrey', borderwidth=1, relief="solid")
cycle_counter_label.grid(row=9, column=2, sticky='nsew', padx=(0,10))

cycle_listbox = tk.Listbox(window)
cycle_listbox.grid(row=5, column=0, sticky='nw', padx=(10, 0))

graph_button = tk.Button(width=10, text="Show Graph", bg='lightgray', command=show_graph)
graph_button.grid(row=5, column=1, sticky='nw')

flag_button = tk.Button(window, text="Flag Cycle", bg='lightgray', command=flag_cycle)
flag_button.grid(row=5, column=1, sticky='nw', pady=30)

flagged_cycles = set()

# Start the GUI event loop
window.mainloop()